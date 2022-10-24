# TODO 1 - Open and read the cells of an Excel document with openpyxl module
import pprint
from openpyxl import load_workbook

print("Opening Excel Spreedsheet...")
spreedSheet = load_workbook("transactions.xlsx")
sheet = spreedSheet.active
supplier_data = {}

print("Reading the rows form spreedsheet...")
for row_Num in range(2, sheet.max_row+1):
    transaction_type = sheet["B"+str(row_Num)].value
    supplier_name = sheet["C"+str(row_Num)].value
    order_amount = sheet["D"+str(row_Num)].value
    
    
    #TODO 2 - Calculate all the transaction amounts and store it in dictionary data structure
    
    #This is to make sure that the key for transaction type exists
    supplier_data.setdefault(transaction_type, {})
    
    #This is to make sure that the supplier exists in the dictinary
    supplier_data[transaction_type].setdefault(supplier_name, {"transaction_count": 0,
                                                               "amount": 0})
    
    supplier_data[transaction_type][supplier_name]["transaction_count"] += 1
    
    supplier_data[transaction_type][supplier_name]["amount"] += int(order_amount)


#TODO Write the data structure to .py file using pprint module

print("Writing to the outpu file...")
with open("output.py", "w") as output_file:
    output_file.write("All transactions = "+ pprint.pformat(supplier_data))
print("Done writing the output file")    